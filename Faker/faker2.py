from faker import Faker
from openpyxl import Workbook

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Fake Data"

# Create Faker object
fake_data = Faker()

# Headers
ws.cell(row=1, column=1).value = "Name"
ws.cell(row=1, column=2).value = "Email"
ws.cell(row=1, column=3).value = "Address"
ws.cell(row=1, column=4).value = "Phone Number"

# Generate 80 records
for i in range(2, 82):   # Rows 2 to 81 = 80 records
    ws.cell(row=i, column=1).value = fake_data.name()
    ws.cell(row=i, column=2).value = fake_data.email()
    ws.cell(row=i, column=3).value = fake_data.address()
    ws.cell(row=i, column=4).value = fake_data.phone_number()

    print(f"Record {i-1}: {ws.cell(row=i, column=1).value}")

# Save Excel file
wb.save("fAke_data.xlsx")

print("Successfully created fake_data.xlsx with 80 records.")